#!/usr/bin/env python3
"""
TechyEra WhatsApp Sender
Reads phone numbers from the resume scanner Excel and sends
personalized WhatsApp messages via WhatsApp Web.

BEFORE RUNNING:
  1. Open WhatsApp Web in Chrome: https://web.whatsapp.com
  2. Scan the QR code with your phone
  3. Keep the browser open while this script runs
"""

import re
import time
import random
import pywhatkit
import openpyxl
from pathlib import Path
from datetime import datetime

# ── Configuration ──────────────────────────────────────────────────────────────

SEND_MODE       = False     # ← Keep False to preview. Set True only when ready.
TEST_NUMBER     = ""        # ← Fill this for single test e.g. "+917987460954"
DELAY_MIN       = 17        # Minimum delay between messages (seconds)
DELAY_MAX       = 23        # Maximum delay between messages (seconds)
WAIT_TIME       = 12        # Seconds pywhatkit waits before sending each message
TAB_CLOSE       = True      # Close browser tab after each message
DEFAULT_COUNTRY = "+91"     # Used when number has no country code (change to +1 for US)

# ── WhatsApp Message Template ──────────────────────────────────────────────────
# Keep it conversational — WhatsApp is not email.
# Use *text* for bold in WhatsApp.

MESSAGE = """Hiii! 👋

It was great connecting with you!

As discussed, I'm available to support you with your upcoming *interviews and online assessments*.

✅ *Interview Proxy Support* — Any IT role, any company, any round
✅ *Online Assessments* — Coding tests & technical screenings
✅ *Flexible Scheduling* — Available as per your interview slot
✅ *End-to-End Support* — From first round to final round

If you have an interview or assessment coming up, please reach out to me *as early as possible* so we can align on availability before your slot! 🙌

🤝 *Know someone with an upcoming interview?*
Please refer your friends, colleagues, or consultancy contacts who need support with their IT interviews or assessments. Your referral means a lot!

*Mohammad / Kumar*
TechyEra — Interview & Assessment Support
🌐 techyera.co
📧 techyerainterview@gmail.com"""

# ── Phone number cleaner ───────────────────────────────────────────────────────

def clean_phone(raw: str) -> str | None:
    """
    Normalize a phone number to E.164 format (+<countrycode><number>).
    Returns None if the number looks invalid.
    """
    if not raw:
        return None

    # Strip everything except digits and leading +
    digits_only = re.sub(r"[^\d+]", "", str(raw).strip())

    # Already has country code
    if digits_only.startswith("+"):
        num = re.sub(r"\D", "", digits_only)
        if 10 <= len(num) <= 13:
            return "+" + num
        return None

    digits = re.sub(r"\D", "", digits_only)

    # Indian mobile: 10 digits starting with 6-9
    if len(digits) == 10 and digits[0] in "6789":
        return "+91" + digits

    # US/Canada: 10 digits starting with 2-9
    if len(digits) == 10 and digits[0] in "23456789":
        return "+1" + digits

    # Already has country code digits (no +)
    if 11 <= len(digits) <= 13:
        return "+" + digits

    return None


# ── Excel loader ───────────────────────────────────────────────────────────────

def find_latest_excel() -> Path | None:
    downloads = Path.home() / "Downloads"
    files = sorted(
        downloads.glob("TechyEra_Candidates_*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True
    )
    return files[0] if files else None


def load_numbers(excel_path: Path) -> tuple[list[dict], any, any, any]:
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Candidates"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    seen, candidates = set(), []

    for row in range(2, ws.max_row + 1):
        phone  = str(ws.cell(row, col("Phone")).value or "").strip()
        name   = ws.cell(row, col("Candidate Name")).value or "Unknown"
        status = ws.cell(row, col("Status")).value or ""

        if status == "WA Sent":        # skip already messaged
            continue

        clean = clean_phone(phone)
        if not clean:
            continue
        if clean in seen:              # deduplicate
            continue
        seen.add(clean)

        candidates.append({
            "row":   row,
            "name":  name,
            "phone": clean,
        })

    return candidates, wb, ws, col


# ── Sender ─────────────────────────────────────────────────────────────────────

def send_single(phone: str) -> bool:
    """Send to one number. Returns True on success."""
    try:
        pywhatkit.sendwhatmsg_instantly(
            phone_no      = phone,
            message       = MESSAGE,
            wait_time     = WAIT_TIME,
            tab_close     = TAB_CLOSE,
            close_time    = 3,
        )
        return True
    except Exception as e:
        print(f"    ERROR: {e}")
        return False


# ── Dry run preview ────────────────────────────────────────────────────────────

def dry_run(candidates: list[dict]):
    print(f"\n{'─'*60}")
    print(f"  DRY RUN — WhatsApp Web will NOT open")
    print(f"  Total numbers loaded: {len(candidates)}")
    print(f"{'─'*60}")
    print(f"\n  First 5 numbers that will receive the message:\n")
    for i, c in enumerate(candidates[:5], 1):
        print(f"  {i}. {c['name']:<35} {c['phone']}")
    if len(candidates) > 5:
        print(f"\n  ... and {len(candidates) - 5} more.")
    print(f"\n  MESSAGE PREVIEW:")
    print(f"{'─'*60}")
    print(MESSAGE)
    print(f"{'─'*60}")
    print(f"\n  To send to a single test number:")
    print(f"    1. Set TEST_NUMBER = '+91XXXXXXXXXX'")
    print(f"    2. Set SEND_MODE   = True")
    print(f"    3. Run again\n")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  TechyEra WhatsApp Sender")
    print("=" * 60)

    # ── Single test mode ───────────────────────────────────────────
    if SEND_MODE and TEST_NUMBER:
        clean = clean_phone(TEST_NUMBER)
        if not clean:
            print(f"  ❌ Invalid test number: {TEST_NUMBER}")
            raise SystemExit(1)

        print(f"\n  TEST MODE — Sending to {clean} only")
        print(f"  Make sure WhatsApp Web is open in Chrome!\n")
        time.sleep(3)

        ok = send_single(clean)
        if ok:
            print(f"\n  ✅ Test message sent to {clean}")
        else:
            print(f"\n  ❌ Failed to send to {clean}")
        raise SystemExit(0)

    # ── Load from Excel ────────────────────────────────────────────
    excel_path = find_latest_excel()
    if not excel_path:
        print("  ❌ No TechyEra_Candidates_*.xlsx found in Downloads.")
        print("  Run resume_scanner.py first.")
        raise SystemExit(1)

    print(f"  Excel : {excel_path.name}")
    candidates, wb, ws, col_fn = load_numbers(excel_path)
    print(f"  Valid phone numbers found: {len(candidates)}")

    if not candidates:
        print("  Nothing to send.")
        raise SystemExit(0)

    # ── Dry run ────────────────────────────────────────────────────
    if not SEND_MODE:
        dry_run(candidates)
        raise SystemExit(0)

    # ── Full send ──────────────────────────────────────────────────
    print(f"\n  Make sure WhatsApp Web is open in Chrome!")
    print(f"  Sending to {len(candidates)} numbers with {DELAY_SECONDS}s gap each.")
    confirm = input("\n  Type YES to start: ").strip()
    if confirm != "YES":
        print("  Aborted.")
        raise SystemExit(0)

    total  = len(candidates)
    sent   = 0
    failed = 0

    for i, c in enumerate(candidates, 1):
        print(f"  [{i:>4}/{total}] Sending → {c['name']:<30} {c['phone']} ...", end=" ")
        ok = send_single(c["phone"])

        if ok:
            sent += 1
            ws.cell(c["row"], col_fn("Status")).value = "WA Sent"
            print("✅")
        else:
            failed += 1
            print("❌")

        if i < total:
            delay = random.randint(DELAY_MIN, DELAY_MAX)
            print(f"         ⏳ Waiting {delay}s before next message...")
            time.sleep(delay)

    wb.save(str(excel_path))

    print(f"\n{'='*60}")
    print(f"  ✅ Sent   : {sent}")
    print(f"  ❌ Failed : {failed}")
    print(f"  Excel updated — sent marked as 'WA Sent'")
    print(f"{'='*60}")
