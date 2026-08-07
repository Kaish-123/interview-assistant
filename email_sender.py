#!/usr/bin/env python3
"""
TechyEra Gmail Sender
Reads candidate emails from the resume scanner Excel and sends personalized outreach.
Run in DRY RUN mode first to preview — no emails sent until you set SEND_MODE = True.
"""

import smtplib
import time
import re
from pathlib import Path
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import openpyxl

# ── Configuration — FILL THESE BEFORE RUNNING ─────────────────────────────────

SENDER_EMAIL    = "techyerainterview@gmail.com"
SENDER_PASSWORD = "nzuy gmsp kqna rhzc"
SENDER_NAME     = "Mohammad / Kumar | TechyEra"

SEND_MODE       = False         # ← Keep False to preview. Set True only when ready to send.
DELAY_SECONDS   = 8             # Wait between emails (Gmail safe: 500/day max)
MAX_EMAILS      = 500           # Hard cap per run (Gmail daily limit)

# ── Email Template ─────────────────────────────────────────────────────────────
# {name}      → candidate's name
# {role}      → their current role (or "IT Professional" if unknown)
# {sender}    → your name

SUBJECT = "Following Up — Interview & Assessment Support | TechyEra"

BODY_HTML = """\
<html>
<body style="font-family: Arial, sans-serif; font-size: 14px; color: #222;
             max-width: 580px; margin: 0 auto; line-height: 1.6;">

  <!-- Header -->
  <div style="background: #1a1a2e; padding: 22px 28px; border-radius: 8px 8px 0 0;">
    <h2 style="margin: 0; color: #ffffff; font-size: 20px; letter-spacing: 0.5px;">
      TechyEra <span style="color: #4fc3f7;">Interview Support</span>
    </h2>
    <p style="margin: 4px 0 0; color: #90caf9; font-size: 12px;">
      Interview Coaching &nbsp;|&nbsp; Proxy Support &nbsp;|&nbsp; Online Assessments
    </p>
  </div>

  <!-- Body -->
  <div style="background: #f9f9f9; padding: 28px; border: 1px solid #e0e0e0;
              border-top: none; border-radius: 0 0 8px 8px;">

    <p>Hiii,</p>

    <p>It was great connecting with you!</p>

    <p>
      As discussed, I wanted to follow up and remind you that I am available
      to support you with your upcoming
      <strong>interviews and online assessments</strong>.
    </p>

    <!-- WhatsApp CTA - visible early -->
    <div style="background: #e8f5e9; border-left: 4px solid #2e7d32;
                padding: 12px 18px; border-radius: 4px; margin: 18px 0;">
      <p style="margin: 0; color: #2e7d32; font-weight: bold; font-size: 14px;">
        &#x1F4F2; &nbsp;Reach me directly on WhatsApp:
        <a href="https://wa.me/917987460954" style="color: #1b5e20; text-decoration: none;">
          +91 7987460954
        </a>
      </p>
    </div>

    <!-- Service box -->
    <div style="background: #e8f4fd; border-left: 4px solid #1565c0;
                padding: 14px 18px; border-radius: 4px; margin: 18px 0;">
      <p style="margin: 0 0 8px; font-weight: bold; color: #1565c0;">
        What I can help you with:
      </p>
      <ul style="margin: 0; padding-left: 18px; color: #333;">
        <li>&#x2705; &nbsp;<strong>Interview Proxy Support</strong> — Any IT role, any company, any round</li>
        <li>&#x2705; &nbsp;<strong>Online Assessments</strong> — Coding tests &amp; technical screenings</li>
        <li>&#x2705; &nbsp;<strong>Flexible Scheduling</strong> — Available as per your interview slot</li>
        <li>&#x2705; &nbsp;<strong>End-to-End Support</strong> — From first round to final round</li>
      </ul>
    </div>

    <p>
      If you have an interview or assessment coming up, please reach out to me
      <strong>as early as possible</strong> so we can prepare and align on availability
      before your scheduled slot.
    </p>

    <!-- Referral box -->
    <div style="background: #fff8e1; border-left: 4px solid #f9a825;
                padding: 14px 18px; border-radius: 4px; margin: 18px 0;">
      <p style="margin: 0; color: #555;">
        &#x1F91D; &nbsp;<strong>Know someone with an upcoming interview?</strong><br>
        Please refer your friends, colleagues, or consultancy contacts who need
        support with their <strong>IT interviews or online assessments</strong>.
        Your referral means a lot!
      </p>
    </div>

    <p>
      Simply reply to this email or ping me on WhatsApp — I'll get back to you promptly.
    </p>

    <p>Looking forward to supporting you!</p>

    <!-- Signature -->
    <div style="border-top: 1px solid #ddd; margin-top: 24px; padding-top: 16px;">
      <p style="margin: 0; font-weight: bold; color: #1a1a2e; font-size: 15px;">
        Mohammad / Kumar
      </p>
      <p style="margin: 6px 0 0; color: #555; font-size: 13px; line-height: 1.8;">
        TechyEra — Interview &amp; Assessment Support<br>
        &#x1F4F2; WhatsApp: <a href="https://wa.me/917987460954"
          style="color: #1565c0; text-decoration: none;">+91 7987460954</a><br>
        &#x1F4E7; <a href="mailto:techyerainterview@gmail.com"
          style="color: #1565c0; text-decoration: none;">techyerainterview@gmail.com</a><br>
        &#x1F310; <a href="https://techyera.co/"
          style="color: #1565c0; text-decoration: none;">techyera.co</a>
      </p>
    </div>

    <p style="font-size: 11px; color: #aaa; margin-top: 20px;">
      To unsubscribe from future emails, simply reply with "Unsubscribe".
    </p>

  </div>
</body>
</html>
"""

BODY_PLAIN = """\
Hiii,

It was great connecting with you!

As discussed, I wanted to follow up and remind you that I am available
to support you with your upcoming interviews and online assessments.

📲 Reach me directly on WhatsApp: +91 7987460954

WHAT I CAN HELP YOU WITH:
  ✅ Interview Proxy Support — Any IT role, any company, any round
  ✅ Online Assessments — Coding tests & technical screenings
  ✅ Flexible Scheduling — Available as per your interview slot
  ✅ End-to-End Support — From first round to final round

If you have an interview or assessment coming up, please reach out to me
as early as possible so we can prepare and align on availability before
your scheduled slot.

REFERRAL REQUEST:
Know someone with an upcoming interview? Please refer your friends,
colleagues, or consultancy contacts who need support with their IT
interviews or online assessments. Your referral means a lot!

Simply reply to this email or ping me on WhatsApp — I'll get back
to you promptly.

Looking forward to supporting you!

Mohammad / Kumar
TechyEra — Interview & Assessment Support
📲 WhatsApp: +91 7987460954
📧 techyerainterview@gmail.com
🌐 techyera.co

---
To unsubscribe, reply with "Unsubscribe".
TechyEra — IT Staffing & Interview Coaching

---
To unsubscribe, reply with "Unsubscribe".
"""

# ── Excel reader ───────────────────────────────────────────────────────────────

def find_latest_excel() -> Path | None:
    downloads = Path.home() / "Downloads"
    files = sorted(
        downloads.glob("TechyEra_Candidates_*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True
    )
    return files[0] if files else None


def _clean_role(role: str) -> str:
    """Return a clean job title or 'IT Professional' if the role looks like a sentence/fragment."""
    role = role.strip()
    if not role:
        return "IT Professional"
    # If too long, has commas, starts lowercase, or is a sentence → generic fallback
    if (
        len(role) > 50
        or "," in role
        or role[0].islower()
        or role.lower().startswith(("a ", "an ", "the ", "end-to-end", "chance"))
        or len(role.split()) > 6
    ):
        return "IT Professional"
    return role


def load_candidates(excel_path: Path) -> list[dict]:
    """Load candidates who have an email and haven't been emailed yet."""
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Candidates"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    seen_emails = set()
    candidates  = []

    for row in range(2, ws.max_row + 1):
        email  = (ws.cell(row, col("Email")).value or "").strip().lower()
        status = ws.cell(row, col("Status")).value or ""
        name   = ws.cell(row, col("Candidate Name")).value or ""
        role   = ws.cell(row, col("Current Role")).value or ""

        if not email or "@" not in email:
            continue
        if status == "Emailed":          # skip already-emailed candidates
            continue
        if email in seen_emails:         # deduplicate by email address
            continue
        seen_emails.add(email)

        candidates.append({
            "row":   row,
            "name":  name if name not in ("", "Unknown") else "",
            "role":  _clean_role(role),
            "email": email,
        })

    return candidates, wb, ws, col


# ── Email builder ──────────────────────────────────────────────────────────────

def build_message(candidate: dict) -> MIMEMultipart:
    # Use first name if known, otherwise a neutral greeting
    raw_name = candidate["name"]
    name = raw_name.split()[0].capitalize() if raw_name else "there"
    role = candidate["role"][:60]

    msg = MIMEMultipart("alternative")
    msg["Subject"] = SUBJECT.format(role=role)
    msg["From"]    = f"{SENDER_NAME} <{SENDER_EMAIL}>"
    msg["To"]      = candidate["email"]

    plain = BODY_PLAIN.format(name=name, role=role, sender=SENDER_NAME)
    html  = BODY_HTML.format(name=name, role=role, sender=SENDER_NAME)

    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html,  "html"))
    return msg


# ── Sender ─────────────────────────────────────────────────────────────────────

def send_emails(candidates: list[dict], wb, ws, col_fn):
    total   = min(len(candidates), MAX_EMAILS)
    sent    = 0
    failed  = 0
    log_path = Path(__file__).parent / "email_log.txt"

    print(f"\n  Connecting to Gmail SMTP...")
    try:
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        print(f"  ✅ Connected as {SENDER_EMAIL}")
    except Exception as e:
        print(f"  ❌ Login failed: {e}")
        print("\n  Make sure you're using a Gmail App Password, not your login password.")
        print("  Generate one at: myaccount.google.com → Security → App Passwords")
        return

    with open(log_path, "a") as log:
        log.write(f"\n=== Run: {datetime.now().strftime('%Y-%m-%d %H:%M')} ===\n")

        for i, candidate in enumerate(candidates[:MAX_EMAILS], 1):
            try:
                msg = build_message(candidate)
                server.sendmail(SENDER_EMAIL, candidate["email"], msg.as_string())

                # Mark as Emailed in Excel
                ws.cell(candidate["row"], col_fn("Status")).value = "Emailed"

                sent += 1
                log.write(f"SENT   | {candidate['email']} | {candidate['name']}\n")
                print(f"  [{i}/{total}] ✅ Sent → {candidate['name']:<25} <{candidate['email']}>")

            except Exception as e:
                failed += 1
                log.write(f"FAILED | {candidate['email']} | {e}\n")
                print(f"  [{i}/{total}] ❌ Failed → {candidate['email']}: {e}")

            if i < total:
                time.sleep(DELAY_SECONDS)

    server.quit()

    # Save updated Excel (Status = Emailed for sent rows)
    wb.save(str(excel_path))
    print(f"\n  Excel updated — sent candidates marked as 'Emailed'.")
    print(f"  Log saved to: {log_path}")

    return sent, failed


# ── Dry run preview ────────────────────────────────────────────────────────────

def dry_run_preview(candidates: list[dict]):
    print(f"\n{'─'*62}")
    print(f"  DRY RUN — No emails will be sent")
    print(f"  Total candidates with email: {len(candidates)}")
    print(f"{'─'*62}")

    # Show first 5 previews
    for i, c in enumerate(candidates[:5], 1):
        name = c["name"].split()[0].capitalize() if c["name"] else "there"
        role = c["role"][:60]
        print(f"\n  Preview {i}:")
        print(f"  To      : {c['name']} <{c['email']}>")
        print(f"  Subject : {SUBJECT.format(role=role)}")
        print(f"  Greeting: Hi {name},")
        print(f"  Role ref: ...your background as a {role}...")

    if len(candidates) > 5:
        print(f"\n  ... and {len(candidates) - 5} more candidates.")

    print(f"\n{'─'*62}")
    print(f"  To actually send:")
    print(f"  1. Set SENDER_EMAIL    = 'your@gmail.com'")
    print(f"  2. Set SENDER_PASSWORD = 'your-16-char-app-password'")
    print(f"  3. Set SEND_MODE       = True")
    print(f"  4. Run again")
    print(f"{'─'*62}\n")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 62)
    print("  TechyEra Gmail Sender")
    print("=" * 62)

    # Find Excel
    excel_path = find_latest_excel()
    if not excel_path:
        print("  ❌ No TechyEra_Candidates_*.xlsx found in Downloads.")
        print("  Run resume_scanner.py first.")
        raise SystemExit(1)

    print(f"  Excel  : {excel_path.name}")

    candidates, wb, ws, col_fn = load_candidates(excel_path)
    print(f"  Candidates with email (not yet emailed): {len(candidates)}")

    if not candidates:
        print("  Nothing to send — no new email addresses found.")
        raise SystemExit(0)

    if not SEND_MODE:
        dry_run_preview(candidates)
    else:
        if not SENDER_EMAIL or not SENDER_PASSWORD:
            print("  ❌ SENDER_EMAIL and SENDER_PASSWORD must be set before sending.")
            raise SystemExit(1)

        print(f"\n  Mode   : LIVE SEND")
        print(f"  From   : {SENDER_EMAIL}")
        print(f"  Cap    : {MAX_EMAILS} emails this run")
        print(f"  Delay  : {DELAY_SECONDS}s between emails")
        confirm = input("\n  Type YES to start sending: ").strip()
        if confirm != "YES":
            print("  Aborted.")
            raise SystemExit(0)

        sent, failed = send_emails(candidates, wb, ws, col_fn)
        print(f"\n{'='*62}")
        print(f"  ✅ Sent    : {sent}")
        print(f"  ❌ Failed  : {failed}")
        print(f"{'='*62}")
